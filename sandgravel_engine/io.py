import yaml
import json
import os
from pathlib import Path
from .models import BalanceResult, MaterialStream, SizeDistribution, EquipmentSelection

_CONFIG_DIR = Path(__file__).parent / "config"


def load_yaml_config(option_name: str) -> dict:
    path = _CONFIG_DIR / f"{option_name}.yaml"
    if not path.exists():
        raise FileNotFoundError(f"Config file not found: {path}")
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def export_to_excel(result: BalanceResult, path: str):
    import openpyxl
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "物料平衡"
    ws1.append(["名称", "吨位(t/h)", ">150", "150-80", "80-40", "40-20", "20-5", "<5"])
    for name, stream in result.streams.items():
        ws1.append([name, round(stream.tonnage, 2)] + [round(v, 3) for v in stream.grading.to_list()])

    ws2 = wb.create_sheet("设备选型")
    ws2.append(["型号", "台数", "单机能力(t/h)", "实际通过量(t/h)", "负荷率"])
    for eq in result.equipment:
        ws2.append([eq.model, eq.quantity, eq.unit_capacity, round(eq.actual_throughput, 2), round(eq.load_factor, 3)])

    ws3 = wb.create_sheet("收敛信息")
    ws3.append(["迭代次数", "收敛误差"])
    ws3.append([result.iterations, result.convergence_error])

    wb.save(path)


def import_from_excel(path: str) -> BalanceResult:
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)

    streams = {}
    if "物料平衡" in wb.sheetnames:
        ws1 = wb["物料平衡"]
        for row in ws1.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            name = str(row[0])
            tonnage = float(row[1]) if row[1] else 0.0
            grading_values = [float(v) if v else 0.0 for v in row[2:8]]
            streams[name] = MaterialStream(name=name, tonnage=tonnage, grading=SizeDistribution.from_list(grading_values))

    equipment = []
    if "设备选型" in wb.sheetnames:
        ws2 = wb["设备选型"]
        for row in ws2.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            equipment.append(EquipmentSelection(
                model=str(row[0]), quantity=int(row[1]), unit_capacity=float(row[2]),
                actual_throughput=float(row[3]) if row[3] else 0.0,
            ))

    iterations = 0
    convergence_error = 0.0
    if "收敛信息" in wb.sheetnames:
        ws3 = wb["收敛信息"]
        row = next(ws3.iter_rows(min_row=2, values_only=True), None)
        if row:
            iterations = int(row[0]) if row[0] else 0
            convergence_error = float(row[1]) if row[1] else 0.0

    return BalanceResult(streams=streams, equipment=equipment, iterations=iterations, convergence_error=convergence_error)


def export_to_json(result: BalanceResult) -> str:
    return json.dumps(result.to_dict(), ensure_ascii=False, indent=2)
